#导入章节，试题信息到数据库
import pymysql
import datetime
import json
import threading
from openpyxl import load_workbook
import xlrd
import re

#设定数据库连接
#conn = pymysql.connect(host="139.196.177.17",user="root",password="root1234",db="bst",use_unicode=True,charset="utf8")
#conn.execute('set max_allowed_packet=67108864')

class ImportThread(threading.Thread):
    def __init__(self,name,items,conn,icategory):
        threading.Thread.__init__(self)
        self.name=name
        self.items=items
        self.conn=conn
        self.icategory=icategory
    def run(self):
        print("启动线程:"+self.name+",size:"+str(len(self.items)))
        importExamInfos(self.conn,self.items,self.icategory)

def answers(qanswer,ans):
    str = qanswer
    repl = ("A","B","C","D","E","F","G","H")
    # ans = "B,C"
    anss = re.split(r",",ans)
    qanswers = []
    for item in repl:
        str = re.sub(r""+item+".",item+"."+"1",str)
    print(str)
    for key in anss:
        print(key)
        str = re.sub(r""+key+".1",key+"."+"0",str)
        print(str)

    answers = re.split(r"[A-H].",str.strip())
    print(len(answers))
    index = 1
    for answer in answers:
        if len(answer) != 0:
            answer = re.findall(r"^\S*",answer.strip())
            #print(answer[0])
            answer1 = answer[0][0]
            #print(answer1)
            if answer1 == 0:
                answer1 = answer[0][1,].strip()
            else:
                answer1 = answer[0].strip()
            print(answer1[0:1],answer1[1:])
            qanswers.append({"iscorrect":answer1[0:1],"label":answer1[1:],"img":"","index":index})
            index +=1
    return qanswers

#读取Excel xls
def readExcelInfoFromLocal():
    # 打开文件
    filepath = r'F:\questions\第二章　儿童社会工作.xls'
    category = re.split(r'\\',filepath)[-1]
    category = re.split(r"章",category)[1].strip()
    category = re.split(r"\.",category)[0].strip()
    print(category)
    workbook = xlrd.open_workbook(filepath)
    # 获取所有sheet
    # print(workbook.sheet_names()) # [u'sheet1', u'sheet2']
    #sheet_name = workbook.sheet_names()[0]

    # 根据sheet索引或者名称获取sheet内容
    sheet1 = workbook.sheet_by_index(0)  # sheet索引从0开始
    #sheet2 = workbook.sheet_by_name('sheet2')

    # sheet的名称，行数，列数
    print(sheet1.name, sheet1.nrows, sheet1.ncols)

    exams = []
    row = 0;
    while row < sheet1.nrows:
        try:
            qyname = sheet1.cell(row, 0).value   # 题目标题
            qanswers = sheet1.cell(row, 1).value # 题目设定的4个答案,多选5个
            qcorrect = sheet1.cell(row, 2).value  # 题目的正确答案
            qcontent = sheet1.cell(row, 3).value # 题目解析
            qcontent = qcontent.lstrip()
            # qcontent = re.findall(r"\S*",qcontent)[0].strip()
            qcontent = re.split(r"考前内部押题",qcontent)[0].strip()
            type = re.findall(r"\[\S*\]",qyname)[0]

            qyname = re.split(r"\[\S*\]",qyname)[1].strip()
            qyname = re.sub(r"\xa0",'',qyname)

            # qanswers = re.findall(r"\S*",qanswers)[0]

            # 针对所有的答案，进行筛选，单选有4个，多选可能会有5个或以上的答案
            qanswers = answers(qanswers, qcorrect)

            # temp = re.split(r"B.",qanswers)[1]
            # qanswer_a = re.split(r"B.",qanswers)[0].strip()
            # qanswer_a = re.split(r"A.",qanswer_a)[1].strip()
            # qanswer_b = re.split(r"C.",temp)[0].strip()
            # temp = re.split(r"C.",temp)[1]
            # qanswer_c = re.split(r"D.",temp)[0].strip()
            # temp = re.split(r"D.", temp)[1]
            # qanswer_d = re.split(r"D.",temp)[0].strip()

            if re.match(r"\S*单选", type):
                type = 0  #单选
            elif re.match(r"\S*多选", type):
                type = 1  #多选
            elif re.match(r"\S*简答", type):# 简单
                type = 3
            exams.append({"type":type,"qyname":qyname,"qanswers":qanswers,"qcorrect":qcorrect,"qcontent":qcontent,"titleImg":"","jxImg":"","category":category,"icategory":""})
        except Exception as e:
            print("读取Excel中题目信息")
            print(e)
            return
        row += 1
    # print(str(exams))
    return exams

def threadImportInfo(items,listSize=20):
    examSize = len(exams)
    if listSize <= 0 :
        listSize = 20

    if listSize >= examSize :
        listSize = examSize

    threadSize = examSize//listSize
    yushu = examSize%listSize

    if yushu != 0 :
        threadSize = threadSize+1

    #创建线程
    threadLock = threading.Lock()
    threads = []
    for i in range(1,threadSize+1):
        conn = pymysql.connect(host="139.196.177.17",user="root",password="root1234",db="bst",use_unicode=True,charset="utf8")
        # icategory = addCategory(conn,items)
        icategory = 183
        threadItem = ImportThread("ImportThread:"+str(i),items[(i-1)*listSize:i*listSize],conn,icategory)
        threadItem.start()
        threads.append(threadItem)

    # 等待所有线程完成
    for t in threads:
        t.join()

def addCategory(conn,items):
    cur = conn.cursor()
    sql = "INSERT INTO page_category (cname,ipid,ENABLE,igrade,isubjectid,icreator,ccreator,dcreator,imodify,cmodify,dmodify) VALUES ('"+str(items[0]['category'])+"',70,0,5,70,1,'admin',now(),1,'admin',now())"
    cur.execute(sql)
    conn.commit()
    cur.close()
    return cur.lastrowid

def importExamInfos(conn,items,icategory):
    # icategory = addCategory(conn, items)
    for item in items:
        item['icategory'] = icategory
        importDB(conn,item)

def importDB(conn,item):
    # pagerId = addPager(conn,item)
    # questions = readExamQuestionInfo(item["sjname"])
    # print("试卷:"+item["sjname"])
    # for question in item:
    questionId = addQuestion(conn,item)
    #addExamQuextion(conn,pagerId,questionId,question["qtype"],question["xh"])
    addAnswer(conn,item["qanswers"],questionId)


#添加试卷
# def addPager(conn,examItem):
#     cur = conn.cursor()
#     sql = "insert into pager (ctitle,page_category_id,ccontent,icreator, ccreator, dcreator, imodify, cmodify, dmodify,itotaltime, isinglenumber, imultinumber, ijudgenumber, iessaynumber,itotalscore,isinglescore, imultiscore, ijudgescore, iessayscore) values (%s,3,%s,20,'admin',now(),20,'admin',now(),120,%s,0,0,0,%s,0,0,0,0) "
#     cur.executemany(sql,[(examItem["sjname"],examItem["sjid"],examItem["tmsl"],examItem["tmsl"])])
#     conn.commit()
#     cur.close()
#     autoId = cur.lastrowid;
#     #获取自增id
#     return autoId

#获取试卷对应的题目
# def readExamQuestionInfo(examName):
#     filePath = "./questions/"+examName+".xlsx"
#     wb = load_workbook(filePath)
#     sheet = wb.active
#     questions=[]
#     for row in sheet.rows:
#         xh = row[0].value
#         title = row[2].value
#         titleImg = row[3].value
#         jx=row[6].value
#         qtype=row[9].value
#         answer=json.loads(row[4].value)
#         (ichapterid,cchaptername) = tyChapter(row[7].value)
#         questions.append({"xh":xh,"title":title,"titleImg":titleImg,"jx":jx,"qtype":qtype,"ichapterid":ichapterid,"cchaptername":cchaptername,"jxImg":"","answer":answer})
#     return questions

#添加题目,并返回题目的id
def addQuestion(conn,item):
    print(item)
    title = item["qyname"]
    jx=item["qcontent"]
    titleImg=item["titleImg"]
    jxImg=item["jxImg"]
    qtype=item["type"]
    ichapterid=item["icategory"]
    cchaptername=item["category"]

    cur = conn.cursor()
    cur.executemany("insert into page_question (ctitle, ccontent, ctype, titleimage, contentimage, page_category_id, icreator, ccreator, dcreator, imodify, cmodify, dmodify, isort, ichapterid, cchaptername, cremark, onelevel, twolevel) VALUES (%s, %s, %s, %s, %s, 68, 20, 'admin', now(), 20, 'admin', now(), 0, %s, %s, 'python', 65, 66)",[(title,jx,qtype,titleImg,jxImg,ichapterid,cchaptername)])
    conn.commit()
    cur.close()
    #获取自增id
    return cur.lastrowid

#添加试卷与题目的关联
def addExamQuextion(conn,ipagerid,iquestionid,iquestiontype,isort):
    cur = conn.cursor()
    cur.executemany("INSERT INTO page_pager_question (ipagerid, iquestionid, iquestiontype, isort, iversion, iscore) VALUES (%s, %s, %s, %s, 0, 0)",[(ipagerid,iquestionid,iquestiontype,isort)])
    conn.commit()
    cur.close()

def addAnswer(conn,answerItems,questionId):
    cur = conn.cursor()
    # index=0 #排序
    for item in answerItems :
        sql="insert into page_answer (iscorrect, ccontent, ctype, icreator, ccreator, dcreator, imodify, cmodify, dmodify, imageurl, page_question_id, isort) VALUES ("+str(item["iscorrect"])+", '"+item["label"]+"', '0', 20, 'admin', now(), 20, 'admin', now(), '"+item["img"]+"', "+str(questionId)+","+str(item["index"])+")"
        cur.execute(sql)
        conn.commit()
        # index=index+1
    cur.close()



if __name__ == '__main__':
    exams = readExcelInfoFromLocal()
    threadImportInfo(exams, 50)